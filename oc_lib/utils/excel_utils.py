import traceback

from openpyxl.cell import MergedCell
from openpyxl.styles import PatternFill, Border, Side


def auto_adjust_column_width(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if isinstance(cell, MergedCell):
                continue
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width


def set_data_sheet(worksheet_data, columns, rows):
    try:
        worksheet_data.append(columns)

        fill = PatternFill(start_color="80c0d2", end_color="80c0d2", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for i in range(len(columns)):
            worksheet_data.cell(row=1, column=i+1).fill = fill
            worksheet_data.cell(row=1, column=i+1).border = thin_border

        for row in rows:
            worksheet_data.append(row)

    except Exception as e:
        traceback.print_exc()
        raise Exception(str(e))
